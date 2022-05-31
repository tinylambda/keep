from openpyxl import load_workbook

if __name__ == "__main__":
    workbook = load_workbook(filename="/Users/felix/Downloads/olympic.questions.xlsx")
    print(workbook.sheetnames)
    print(workbook.active)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
