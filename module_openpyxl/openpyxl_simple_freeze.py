from openpyxl import load_workbook

if __name__ == '__main__':
    workbook = load_workbook(filename='/Users/felix/Downloads/reviews-sample.xlsx')
    print(workbook.sheetnames)
    print('-' * 64)

    sheet = workbook.active
    sheet.freeze_panes = 'C2'
    workbook.save(filename='/Users/felix/Downloads/reviews-sample-frozen.xlsx')

