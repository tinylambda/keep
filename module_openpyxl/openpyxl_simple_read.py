from pprint import pprint
from openpyxl import load_workbook


if __name__ == '__main__':
    workbook = load_workbook('/Users/felix/Downloads/sample.xlsx')
    print(workbook.sheetnames)

    sheet = workbook.active
    print(sheet, sheet.title)

    print(sheet['A1'])
    print(sheet['A1'].value)
    print(sheet['F10'].value)

    print(sheet.cell(row=10, column=6))
    print(sheet.cell(row=10, column=6).value)

    print('-' * 64)
    print('iterate through the data')
    print(sheet['A1:C2'])

    print('-' * 64)
    print('get all cells from column A')
    pprint(sheet['A'])

    print('-' * 64)
    print('get all cells for a range of columns')
    pprint(sheet['A:B'])

    print('-' * 64)
    print('get all cells from row 5')
    pprint(sheet[5])

    print('-' * 64)
    print('get all cells for a range of rows')
    pprint(sheet[5:6])

    print('-' * 64)
    print('iter_rows')
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
        print(row)

    print('-' * 64)
    print('iter_cols')
    for column in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3):
        print(column)

    print('-' * 64)
    print('values_only')
    for value in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
        print(value)

    print('-' * 64)
    print('.rows and .columns is shortcuts to .iter_rows() and .iter_cols() without any arguments')
    for row in sheet.rows:
        print(row)

    print('-' * 64)
    print('get header information')
    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        print(value)

    print('-' * 64)
    print('extract product information')
    for value in sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):
        print(value)

