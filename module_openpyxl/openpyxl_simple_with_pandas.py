import pathlib
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

data = {
    'product name': ['product 1', 'product 2'],
    'sales month 1': [10, 20],
    'sales month 2': [5, 35],
}

df = pd.DataFrame(data)
print(df)

workbook = Workbook()
sheet = workbook.active

for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)


filepath = pathlib.Path.home().joinpath('Downloads/test14.xlsx')
workbook.save(filepath)

workbook = load_workbook(filepath)
sheet = workbook.active
values = sheet.values
df = pd.DataFrame(values)
print(df)



