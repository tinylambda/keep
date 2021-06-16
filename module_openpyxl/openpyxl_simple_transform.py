import json

from openpyxl import load_workbook

if __name__ == '__main__':
    workbook = load_workbook(filename='/Users/felix/Downloads/sample.xlsx')
    sheet = workbook.active
    products = {}

    header = []
    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        header = value[3:7]

    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):
        row_dict = dict(zip(header, row))
        products[row_dict.pop('product_id')] = row_dict

    print(json.dumps(products))

