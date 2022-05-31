from openpyxl import load_workbook

if __name__ == "__main__":
    workbook = load_workbook(filename="/Users/felix/Downloads/sample.xlsx")
    print(workbook.sheetnames)
    print("-" * 64)

    default_sheet = workbook[workbook.sheetnames[0]]
    default_sheet.title = "new sheet name"
    print(workbook.sheetnames)

    print("-" * 64)
    operations_sheet = workbook.create_sheet("operations")
    print(workbook.sheetnames)

    print("-" * 64)
    hr_sheet = workbook.create_sheet("HR", 0)
    print(workbook.sheetnames)

    print("-" * 64)
    workbook.remove(operations_sheet)
    print(workbook.sheetnames)

    print("-" * 64)
    workbook.remove(hr_sheet)
    print(workbook.sheetnames)

    # copy worksheet
    print("-" * 64)
    print(workbook.sheetnames)

    print("-" * 64)
    workbook.copy_worksheet(default_sheet)
    print(workbook.sheetnames)
