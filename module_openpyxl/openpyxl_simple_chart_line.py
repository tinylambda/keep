import pathlib
import random

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


if __name__ == "__main__":
    workbook = Workbook()
    sheet = workbook.create_sheet("Sample Sheet")
    rows = [
        [
            "",
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ],
        [
            1,
        ],
        [
            2,
        ],
        [
            3,
        ],
    ]

    for row in rows:
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=13):
        for cell in row:
            cell.value = random.randrange(5, 100)

    chart = LineChart()
    data = Reference(worksheet=sheet, min_row=2, max_row=4, min_col=1, max_col=13)
    chart.add_data(data, from_rows=True, titles_from_data=True)
    sheet.add_chart(chart, "C6")

    filepath = pathlib.Path.home().joinpath("Downloads/test11.xlsx")
    workbook.save(filepath)
