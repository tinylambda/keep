import pathlib

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

if __name__ == '__main__':
    red_background = PatternFill(bgColor='00FF0000')
    diff_style = DifferentialStyle(fill=red_background)
    rule = Rule(type='expression', dxf=diff_style)
    rule.formula = ['$A1<0']

    workbook = Workbook()
    sheet = workbook.create_sheet('Sample Sheet')
    sheet.append((1, 2, 3))
    sheet.append((1, 2, 3))
    sheet.append((-1, 2, 3))
    sheet.append((-2, 2, 3))
    sheet.append((-2, 2, 3))
    sheet.append((2, 2, 3))
    sheet.conditional_formatting.add('A1:o100', rule)

    filepath = pathlib.Path.home().joinpath('Downloads/test3.xlsx')
    workbook.save(filepath)
