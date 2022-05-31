import dataclasses

from openpyxl import load_workbook

from simple_classes import Product
from simple_classes import Review

if __name__ == "__main__":
    # create Product objects from the data
    product_field_names = [item.name for item in dataclasses.fields(Product)]
    review_field_names = [item.name for item in dataclasses.fields(Review)]

    workbook = load_workbook(filename="/Users/felix/Downloads/sample.xlsx")
    sheet = workbook.active
    header = []
    # the first line is header information
    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        header = [item for item in value]

    products = []
    reviews = []
    # iterate over all rows and create product objects and review objects
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(header, row))

        product_args = []
        for field_name in product_field_names:
            value = next(
                (
                    item
                    for item in map(
                        row_dict.get, [field_name, "_".join(["product", field_name])]
                    )
                    if item is not None
                ),
                None,
            )
            product_args.append(value)
        product = Product(*product_args)
        products.append(product)

        review_args = []
        for field_name in review_field_names:
            if field_name == "stars":
                field_name = "star_rating"

            value = next(
                (
                    item
                    for item in map(
                        row_dict.get, [field_name, "_".join(["review", field_name])]
                    )
                    if item is not None
                ),
                None,
            )
            review_args.append(value)
        review = Review(*review_args)
        reviews.append(review)
    print(products)
    print(reviews)
